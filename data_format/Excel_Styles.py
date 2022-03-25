


from openpyxl.styles import Font, Color , PatternFill , NamedStyle , Alignment


class Excel_Styles():

    # 預設樣式
    default_Style = None ;    # 預設
    error_Style   = None ;    # 錯誤
    align_Center  = Alignment( horizontal='center' , vertical='center' ) # 置中

    def __init__( self ) :
        # 設定 _ 預設樣式
        _default_Style           = NamedStyle(name='default_Style')
        _default_Style.alignment = self.align_Center
        self.default_Style       = _default_Style

        # 設定 _ 錯誤樣式
        _error_Style           = NamedStyle(name='error_Style')
        _error_Style.font      = Font(color='ffffff', bold=True)
        _error_Style.alignment = self.align_Center
        self.error_Style       = _error_Style


    # 設定 _ 預設樣式
    def set_Default_Style( self  , column_List , sheet , idx ):
        for col in column_List :
            index = col + str( idx+1 )
            sheet[ index ].style = self.default_Style

    # 設定 _ 錯誤樣式
    def set_Error_Style( self , column_List , sheet , idx , fg_Color ) :

        self.error_Style.fill = PatternFill( 'solid' , fgColor = fg_Color ) # 設定標籤顏色

        for col in column_List :
            index = col + str( idx+2 )
            sheet[ index ].style = self.error_Style


    # 自訂樣式
    # 錯誤
    def get_Error_Style(self):
        error_Style      = NamedStyle(name="error_Style")
        error_Style.font = Font(color='ffffff', bold=True)
        error_Style.fill = PatternFill("solid", fgColor="FF0000")
        return error_Style




