
'''

   @ 共同操作

    # 剔除：
     * 沒有手機號碼
     * '測試'、'測試員'、'test' ...
     * 重複姓名 ( 僅留名字下有寵物 )
     * 手機門號：改用、停用 ...

'''

from tool.Common import Common
import os
import tool
from openpyxl import load_workbook
from filter_condition.Filter_Customers import Filter_Customers
from data_format.Excel_Styles import Excel_Styles
from openpyxl.styles import Font , Color , PatternFill , NamedStyle , Alignment


# 篩選錯誤條件
filer = Filter_Customers()

# 樣式
style = Excel_Styles()


# 讀取 _ Excel 檔案 ：工作簿、工作表
wb   = load_workbook( 'data_files/客戶.xlsx' )
ws_1 = wb['客戶篩選']                 # 第一張工作表 ( 初始資料 )
ws_2 = wb.create_sheet("經過篩選條件") # 第二張工作表 ( 修改資料 )

# 寫入資料 （ 第二張工作表 ）
column_List  = [ 'A' , 'B' , 'C' , 'D' ]


e_Style           = NamedStyle(name='e_Style')
e_Style.font      = Font(color='ffffff', bold=True)
e_Style.alignment = Alignment(horizontal='center', vertical='center')

def set_Error_Style( e_Style , column_List , sheet , idx , fg_Color ):

    e_Style.fill = PatternFill('solid', fgColor=fg_Color)  # 設定標籤顏色

    for col in column_List:
        index = col + str(idx + 2)
        sheet[index].style = e_Style


# 迭代數值 ( 第一張工作表 )
for index , value in enumerate( ws_1.values ) :

    cus_Name , cus_Id , cus_Phone , has_Pets = (
        value[0] ,  # 客戶姓名
        value[1] ,  # 客戶 master_id
        value[2] ,  # 客戶手機
        value[3]    # 該名字下是否有寵物
    )

    if cus_Name is None : continue

    # @ 篩選條件 :
    # 測試
    is_Name_Has_Test         = '測試' in cus_Name or 'test' in cus_Name
    # 沒有手機號碼
    is_No_MobilePhone        = cus_Phone == ''
    # 手機停用、改號
    is_MobilePhone_Abandoned = '停用' in cus_Name or '不用' in cus_Name or '號' in cus_Name

    # 符合篩選條件 --> 跳過、不輸入資料
    if is_Name_Has_Test or is_No_MobilePhone or is_MobilePhone_Abandoned : continue

    # 輸入資料( 第二張工作表 )
    ws_2.append( [ cus_Name , cus_Id , cus_Phone , has_Pets ] )

    # 列舉清單錯誤 ( 紫色標示 Ex. '先生' , '小姐' , '先生小姐' ....  )
    if filer.is_Error_Customer_Name(cus_Name):
        set_Error_Style( e_Style , column_List , ws_2 , index , 'ca0bec' )

    # 客戶姓名為數字 ( 綠色標示 )
    if cus_Name.isnumeric():
        set_Error_Style( e_Style , column_List , ws_2 , index , '3dd731' )

    # 姓名中有 : 拒接 ( 黑色標示 )
    if '拒接' in cus_Name:
        set_Error_Style( e_Style , column_List , ws_2 , index , '000000' )


# 調整欄位寬度 ( 第二張工作表 )
ws_2.column_dimensions['A'].width = 25
ws_2.column_dimensions['B'].width = 17
ws_2.column_dimensions['C'].width = 15
ws_2.column_dimensions['D'].width = 15


# 存檔
wb.save( 'data_files/客戶_3.xlsx' )
print('存檔成功')

